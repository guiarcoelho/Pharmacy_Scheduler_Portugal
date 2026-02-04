#!/usr/bin/env python3
"""Convert schedule/worker CSVs into a formatted Excel workbook.

This script copies the layout, styles, and conditional formatting from a
user-provided Excel template (e.g., out/Template.xlsx) and injects
fresh data from schedule_<start>_<end>.csv + worker_stats_<start>_<end>.csv.
"""

from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
from pathlib import Path
import re

import openpyxl
import pandas as pd
import yaml
from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import range_boundaries, get_column_letter

_WINDOW_SCHEDULE_RE = re.compile(
    r"^schedule_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})\.csv$"
)




def _extract_special_dates(schedule: pd.DataFrame) -> set:
    if "day_tags" not in schedule.columns:
        return set()
    special_dates = set()
    for date_value, tags in schedule.groupby("date")["day_tags"]:
        if any(str(tag).strip() for tag in tags.dropna()):
            special_dates.add(date_value)
    return special_dates


def _parse_date(raw):
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    return date.fromisoformat(str(raw))


def _load_worker_vacations(path: Path) -> dict[str, set]:
    if not path.exists():
        return {}
    data = yaml.safe_load(path.read_text()) or {}
    workers = data.get("workers", data) or []
    vacations: dict[str, set] = {}

    for worker in workers:
        worker_id = worker.get("id")
        if not worker_id:
            continue
        raw_vacations = worker.get("vacations", []) or []
        if isinstance(raw_vacations, dict):
            raw_vacations = [raw_vacations]
        for p in raw_vacations:
            days = int(p["days"])
            start = _parse_date(p["start"])
            for offset in range(days):
                vacations.setdefault(worker_id, set()).add(start + timedelta(days=offset))

    return vacations

def _load_schedule(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df["date"] = pd.to_datetime(df["date"]).dt.date
    return df


def _find_latest_windowed_schedule(out_dir: Path) -> Path | None:
    candidates = []
    for path in out_dir.glob("schedule_*.csv"):
        if _WINDOW_SCHEDULE_RE.match(path.name):
            candidates.append(path)
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _default_schedule_path(raw_schedule: str | None) -> Path:
    if raw_schedule:
        return Path(raw_schedule)
    latest = _find_latest_windowed_schedule(Path("out"))
    if latest is not None:
        return latest
    return Path("out/schedule.csv")


def _default_stats_path(schedule_path: Path, raw_stats: str | None) -> Path:
    if raw_stats:
        return Path(raw_stats)
    m = _WINDOW_SCHEDULE_RE.match(schedule_path.name)
    if m:
        return schedule_path.with_name(
            f"worker_stats_{m.group(1)}_{m.group(2)}.csv"
        )
    return Path("out/worker_stats.csv")


def _default_excel_path(schedule_path: Path, raw_out: str | None) -> Path:
    if raw_out:
        return Path(raw_out)
    return schedule_path.with_suffix(".xlsx")


def _build_grid(schedule: pd.DataFrame, worker_order: list[str] | None = None) -> pd.DataFrame:
    grid = schedule.pivot(index="worker", columns="date", values="shift").fillna("OFF")
    grid = grid.reindex(sorted(grid.columns), axis=1)
    if worker_order:
        ordered = [w for w in worker_order if w in grid.index]
        remaining = [w for w in grid.index if w not in ordered]
        grid = grid.loc[ordered + remaining]
    else:
        grid = grid.sort_index()
    return grid




def _collect_style_samples(ws):
    weekday_col = None
    weekend_col = None
    service_col = None

    for col in range(2, ws.max_column + 1):
        day_name = ws.cell(1, col).value
        if day_name in ("Saturday", "Sunday") and weekend_col is None:
            weekend_col = col
        if day_name and day_name not in ("Saturday", "Sunday") and weekday_col is None:
            weekday_col = col
        if ws.cell(2, col).fill.patternType and service_col is None:
            service_col = col
        if weekday_col and weekend_col and service_col:
            break

    def style_from(row, col):
        if col is None:
            return None
        cell = ws.cell(row, col)
        return cell._style

    return {
        "weekday_body": style_from(3, weekday_col),
        "weekend_body": style_from(3, weekend_col) or style_from(3, weekday_col),
        "service_date_style": style_from(2, service_col),
    }

def _copy_col_style(ws, src_col: int, dst_col: int) -> None:
    src_dim = ws.column_dimensions[get_column_letter(src_col)]
    dst_dim = ws.column_dimensions[get_column_letter(dst_col)]
    dst_dim.width = src_dim.width
    dst_dim.hidden = src_dim.hidden
    dst_dim.outline_level = src_dim.outline_level
    dst_dim.collapsed = src_dim.collapsed

    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, src_col)
        dst = ws.cell(row, dst_col)
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst._style = src._style
        dst.number_format = src.number_format


def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = src._style
        dst.number_format = src.number_format


def _adjust_conditional_formatting(
    ws,
    old_max_row: int,
    old_max_col: int,
    new_max_row: int,
    new_max_col: int,
) -> None:
    old_rules = dict(ws.conditional_formatting._cf_rules)
    if not old_rules:
        return

    new_rules = {}
    for cf, rules in old_rules.items():
        range_str = str(cf.sqref) if hasattr(cf, "sqref") else str(cf)
        updated_ranges = []
        for part in range_str.split():
            min_col, min_row, max_col, max_row = range_boundaries(part)

            if new_max_row > old_max_row and max_row == old_max_row:
                max_row = new_max_row
            if new_max_col > old_max_col and max_col == old_max_col:
                max_col = new_max_col

            if min_row > new_max_row or min_col > new_max_col:
                continue
            max_row = min(max_row, new_max_row)
            max_col = min(max_col, new_max_col)

            if min_row > max_row or min_col > max_col:
                continue

            updated_ranges.append(
                f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            )

        if not updated_ranges:
            continue

        new_cf = ConditionalFormatting(" ".join(updated_ranges))
        new_rules[new_cf] = rules

    ws.conditional_formatting._cf_rules = new_rules


def _resize_sheet(ws, needed_rows: int, needed_cols: int) -> None:
    old_max_row = ws.max_row
    old_max_col = ws.max_column

    if needed_cols > old_max_col:
        ws.insert_cols(old_max_col + 1, amount=needed_cols - old_max_col)
        for col in range(old_max_col + 1, needed_cols + 1):
            _copy_col_style(ws, old_max_col, col)

    if needed_rows > old_max_row:
        ws.insert_rows(old_max_row + 1, amount=needed_rows - old_max_row)
        for r in range(old_max_row + 1, needed_rows + 1):
            _copy_row_style(ws, old_max_row, r)
    elif needed_rows < old_max_row:
        ws.delete_rows(needed_rows + 1, old_max_row - needed_rows)

    if needed_cols < old_max_col:
        ws.delete_cols(needed_cols + 1, old_max_col - needed_cols)

    _adjust_conditional_formatting(ws, old_max_row, old_max_col, needed_rows, needed_cols)


def _clear_values(ws, max_row: int, max_col: int) -> None:
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _write_grid(
    ws,
    grid: pd.DataFrame,
    special_dates: set,
    style_samples: dict,
    vacations_by_worker: dict[str, set],
) -> None:
    dates = list(grid.columns)
    workers = list(grid.index)

    needed_rows = 2 + len(workers)
    needed_cols = 1 + len(dates)
    _resize_sheet(ws, needed_rows, needed_cols)
    _clear_values(ws, needed_rows, needed_cols)

    ws.cell(1, 1).value = "worker"
    for idx, d in enumerate(dates, start=2):
        ws.cell(1, idx).value = d.strftime("%A")

    for idx, d in enumerate(dates, start=2):
        ws.cell(2, idx).value = datetime.combine(d, datetime.min.time())

    for r_offset, worker in enumerate(workers, start=3):
        ws.cell(r_offset, 1).value = worker
        for c_offset, d in enumerate(dates, start=2):
            ws.cell(r_offset, c_offset).value = grid.loc[worker, d]

    weekday_style = style_samples.get("weekday_body")
    weekend_style = style_samples.get("weekend_body")
    service_style = style_samples.get("service_date_style")
    vac_font = Font(color="FF9C0006", bold=True)
    vac_fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
    for col_idx, d in enumerate(dates, start=2):
        is_weekend = d.weekday() >= 5
        is_service = d in special_dates

        if is_service and service_style is not None:
            date_cell = ws.cell(2, col_idx)
            date_cell._style = service_style
        body_style = weekend_style if is_weekend else weekday_style
        if body_style is not None:
            for row in range(3, needed_rows + 1):
                ws.cell(row, col_idx)._style = body_style

    for row_offset, worker in enumerate(workers, start=3):
        vac_dates = vacations_by_worker.get(worker, set())
        if not vac_dates:
            continue
        for col_idx, d in enumerate(dates, start=2):
            if d in vac_dates:
                cell = ws.cell(row_offset, col_idx)
                cell.value = "VAC"

    vac_rule = CellIsRule(
        operator="equal",
        formula=["\"VAC\""],
        stopIfTrue=False,
        font=vac_font,
        fill=vac_fill,
    )
    vac_start = f"{get_column_letter(2)}3"
    vac_end = f"{get_column_letter(needed_cols)}{needed_rows}"
    ws.conditional_formatting.add(f"{vac_start}:{vac_end}", vac_rule)


def _write_summary(ws, schedule: pd.DataFrame, stats: pd.DataFrame) -> None:
    if schedule.empty:
        min_date = max_date = ""
    else:
        min_date = schedule["date"].min().isoformat()
        max_date = schedule["date"].max().isoformat()

    values = {
        "Total assignments": len(schedule),
        "Date range": f"{min_date} to {max_date}" if min_date else "",
        "Number of workers": len(stats),
        "Total paid hours": f"{stats['total_hours'].sum():.1f}" if not stats.empty else "0.0",
    }

    for row in ws.iter_rows(min_row=1, max_col=2):
        metric = row[0].value
        if metric in values:
            row[1].value = values[metric]


def _write_worker_stats(ws, stats: pd.DataFrame) -> None:
    if stats.empty:
        return

    needed_rows = 1 + len(stats)
    needed_cols = len(stats.columns)
    _resize_sheet(ws, needed_rows, needed_cols)
    _clear_values(ws, needed_rows, needed_cols)

    for c, col in enumerate(stats.columns, start=1):
        ws.cell(1, c).value = col

    for r, (_, row) in enumerate(stats.iterrows(), start=2):
        for c, col in enumerate(stats.columns, start=1):
            ws.cell(r, c).value = row[col]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert schedule CSVs to a formatted Excel workbook using a template."
    )
    parser.add_argument(
        "--schedule",
        default=None,
        help="Path to schedule CSV (default: latest out/schedule_<start>_<end>.csv)",
    )
    parser.add_argument(
        "--stats",
        default=None,
        help="Path to worker stats CSV (default: matching worker_stats_<start>_<end>.csv)",
    )
    parser.add_argument(
        "--template",
        default="out/Template.xlsx",
        help="Path to formatted Excel template (default: out/Template.xlsx)",
    )
    parser.add_argument(
        "--workers",
        default="config/scenarios/pharmacy_pt/workers.yaml",
        help="Path to workers.yaml (default: config/scenarios/pharmacy_pt/workers.yaml)",
    )
    parser.add_argument(
        "--out",
        default=None,
        help="Output Excel path (default: same basename as --schedule with .xlsx)",
    )
    args = parser.parse_args()

    schedule_path = _default_schedule_path(args.schedule)
    stats_path = _default_stats_path(schedule_path, args.stats)
    template_path = Path(args.template)
    workers_path = Path(args.workers)
    out_path = _default_excel_path(schedule_path, args.out)

    if not schedule_path.exists():
        raise FileNotFoundError(f"Schedule CSV not found: {schedule_path}")
    if not stats_path.exists():
        raise FileNotFoundError(f"Worker stats CSV not found: {stats_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template Excel not found: {template_path}")
    if not workers_path.exists():
        raise FileNotFoundError(f"Workers file not found: {workers_path}")

    schedule = _load_schedule(schedule_path)
    stats = pd.read_csv(stats_path)
    worker_order = stats["worker"].tolist() if "worker" in stats.columns else None
    grid = _build_grid(schedule, worker_order)
    special_dates = _extract_special_dates(schedule)
    vacations_by_worker = _load_worker_vacations(workers_path)

    wb = openpyxl.load_workbook(template_path)

    if "Grid" in wb.sheetnames:
        style_samples = _collect_style_samples(wb["Grid"])
        _write_grid(wb["Grid"], grid, special_dates, style_samples, vacations_by_worker)
    if "Summary" in wb.sheetnames:
        _write_summary(wb["Summary"], schedule, stats)
    if "Worker Statistics" in wb.sheetnames:
        _write_worker_stats(wb["Worker Statistics"], stats)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote formatted workbook to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
